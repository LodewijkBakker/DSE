import openpyxl
import numpy as np
import matplotlib.pyplot as plt


def extract_Q_data(filename) -> tuple[dict, np.ndarray]:
    """
    Converts from ESATAN heat flux data (excel file) to a dictionary of Q values for the thermal simulation
    """

    wb = openpyxl.load_workbook(filename)
    ws = wb['radiation_report']
    Q_ESA = {}
    Q_TCS = {}
    times = np.array([int(x[0].value) for x in ws['B'+str(8):'B'+str(38)]])

    for i in range(1, 693):
        Q_ESA[i] = np.array([x[0].value for x in ws['C'+str(38*i-30):'C'+str(38*i)]])

    DST_inst_box = range(1,7)
    batteries = range(7, 13)
    Nadir = range(13, 61)
    propulsion = range(61, 67)
    OBC = range(67, 73)
    sun_shade = range(73, 121)
    DST_baffle = range(121, 161)
    prop_tank_1 = range(161, 167)
    prop_tank_2 = range(167, 173)
    Zenith = range(173, 243)
    Radiator = range(243, 291)
    North = range(291, 339)
    East = range(339, 387)
    South = range(387, 435)
    West = range(435, 483)
    solar_panel_3 = range(483, 553)
    solar_panel_2 = range(553, 623)
    solar_panel_1 = range(623, 693)

    components = [North, East, South, West, Zenith, Nadir, batteries, solar_panel_1, OBC, prop_tank_1, Radiator, DST_inst_box, DST_baffle, solar_panel_2, solar_panel_3, West]
    component_names = ['North', 'East', 'South', 'West', 'Zenith', 'Nadir', 'batteries', 'solar_panel_1', 'OBC', 'prop_tank_1', 'Radiator', 'DST_inst_box', 'DST_baffle', 'solar_panel_2', 'solar_panel_3', 'propulsion']

    for i, comp in enumerate(components):
        Q_TCS[component_names[i]] = np.mean([Q_ESA[x] for x in comp], axis=0)
        if component_names[i] == 'propulsion':
            Q_TCS[component_names[i]] *= 0.8

    return Q_TCS, times


def interpolate_points(Q: dict, times: np.ndarray, orbit_time: int, n: int) -> dict:
    """
    Interpolates the data points to 1 second intervals
    """

    Q_new = {}
    for key in Q.keys():
        Q_new[key] = np.interp(np.arange(0, orbit_time, 1), times, Q[key])
        Q_new[key] = np.tile(Q_new[key], n)

    return Q_new
